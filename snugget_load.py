import os
import csv

import openpyxl # library to read .xlsx format
import psycopg2

# import settings here because we'll follow whatever's set there
# for the database parameters
import disasterinfosite.settings as settings

def main():
  appName = "disasterinfosite"
  appDir = "disasterinfosite"
  dataDir = os.path.join(appDir, "data")
  snuggetFile = os.path.join(dataDir, "snuggets.xlsx")
  overwriteAll = False
  optionalFields = ['intensity', 'image', 'lookup_value', ''] # all other fields in snuggetFile are required. The empty string is to deal with Excel's charming habit of putting a blank column after all data in a CSV.

  dbURL = settings.DATABASES['default']
  dbHost = dbURL['HOST']
  dbPort = dbURL['PORT']
  dbUser = dbURL['USER']
  dbPass = dbURL['PASSWORD']
  dbName = dbURL['NAME']

  with psycopg2.connect(host=dbHost, port=dbPort, user=dbUser, password=dbPass, database=dbName) as conn:
    with conn.cursor() as cur:
      newSnuggets = XLSXDictReader(snuggetFile)
      rowCount = 1 # row 1 consists of field names, so row 2 is the first data row. We'll increment this before first referencing it.
      for row in newSnuggets:
        rowCount += 1
        if allRequiredFieldsPresent(row, optionalFields, rowCount):
          overwriteAll = processRow(appName, snuggetFile, cur, overwriteAll, row)
  print("Snugget load complete. Processed", rowCount, "rows in", snuggetFile)





def allRequiredFieldsPresent(row, optionalFields, rowCount):
  if any(a != '' for a in row.values()): # if the entire row is not empty
    blanks = []
    for key in row.keys():
      if (key not in optionalFields) and (row[key] == ''):
        blanks.append(key)
    if blanks == []:
      return True
    else:
      print("Unable to process row", rowCount, "with content:")
      print(row)
      if len(blanks) > 1:
        print("Because required fields", blanks, "are empty.")
      else:
        print("Because required field", "'" + blanks[0] + "'", "is empty.")
      return False
  else: # the entire row is blank
    print("Skipping empty row", rowCount)
    return False




def processRow(appName, snuggetFile, cur, overwriteAll, row):
  filterColumn = row["shapefile"] + "_filter_id"
  sectionID = getSectionID(appName, row["section"], cur, subsection=False)
  subsectionID = getSectionID(appName, row["subsection"], cur, subsection=True)

  # check if a snugget for this data already exists
  # if we have a lookup value then deal with this value specifically:
  if row["lookup_value"] is not '':  # if it is blank, we'll treat it as matching all existing values
    filterIDs = [findFilterID(appName, row["shapefile"], row["lookup_value"], cur)]
    if filterIDs[0] is None:
      print("Skipping row:")
      print(row)
      print("Because no filter for lookup_value", row["lookup_value"], "was found in", row["shapefile"])
      return overwriteAll
    else:
      oldSnugget = checkForSnugget(appName, sectionID, subsectionID, filterColumn, filterIDs[0], cur)
      overwriteAll = askUserAboutOverwriting(row, oldSnugget, [], snuggetFile, overwriteAll)
  else:
    filterIDs = findAllFilterIDs(appName, row["shapefile"], cur)
    oldSnuggets = []
    for filterID in filterIDs:
      oldSnugget = checkForSnugget(appName, sectionID, subsectionID, filterColumn, filterID, cur)
      if oldSnugget is not None and oldSnugget not in oldSnuggets:
        oldSnuggets.append(oldSnugget)
    overwriteAll = askUserAboutOverwriting(row, None, oldSnuggets, snuggetFile, overwriteAll)

  for filterID in filterIDs:
    removeOldSnugget(appName, sectionID, subsectionID, filterColumn, filterID, cur)
    addTextSnugget(appName, row, sectionID, subsectionID, filterColumn, filterID, cur)

  return overwriteAll

def _get_i18n_db_column(column_name):
  if column_name.startswith('text-'):
    return 'content_' + column_name.split('-')[1]

def addTextSnugget(appName, row, sectionID, subsectionID, filterColumn, filterID, cur):
#   "intensity" -> disasterinfosite_textsnugget.percentage (numeric, null as null)
#   "image" -> disasterinfosite_textsnugget.image
#   "text" -> disasterinfosite_textsnugget.content
  if row["intensity"] == '':
    row["intensity"] = None
  groupID = getGroupID(appName, row["shapefile"], cur)
  cur.execute(
    'INSERT INTO ' + appName + '_snugget (section_id, sub_section_id, group_id, "' + filterColumn + '") VALUES (%s, %s, %s, %s);',
    (str(sectionID), str(subsectionID), str(groupID), str(filterID))
  )
  snuggetID = getSnuggetID(appName, sectionID, subsectionID, filterColumn, filterID, cur);

  # dynamically build list of multilingual content
  i18n_columns = ", ".join([c for c in (_get_i18n_db_column(column_name) for column_name in row.keys()) if c is not None])
  i18n_placeholders = ", ".join(["%s" for c in (_get_i18n_db_column(column_name) for column_name in row.keys()) if c is not None])
  values_to_insert = (snuggetID, row["text-en"])
  for col in i18n_columns.split(", "):
  	values_to_insert = values_to_insert + (row[col.replace("content_", "text-")],)
  values_to_insert = values_to_insert + (row["image"], row["intensity"])

#  cur.mogrify(
#    'INSERT INTO ' + appName + '_textsnugget (snugget_ptr_id, content, ' + i18n_columns + ', image, percentage) VALUES (%s, %s, ' + i18n_placeholders + ', %s, %s);',
#    values_to_insert
#  )

  cur.execute(
    'INSERT INTO ' + appName + '_textsnugget (snugget_ptr_id, content, ' + i18n_columns + ', image, percentage) VALUES (%s, %s, ' + i18n_placeholders + ', %s, %s);',
    values_to_insert
  )
  # For extra credit, set the group's display_name to the heading value.


def readColumnsFrom(appName, table, cur):
  cols = []
  cur.execute(
    "SELECT column_name FROM information_schema.columns WHERE table_name = %s;",
    [appName + "_" + table.lower()]
  )
  for row in cur.fetchall():
    cols.append(row[0])
  return cols





def getSectionID(appName, sectionName, cur, subsection=False):
  if subsection:
    tableName = appName + "_snuggetsubsection"
  else:
    tableName = appName + "_snuggetsection"

  cur.execute("SELECT MIN(id) FROM " + tableName + " WHERE name = %s;", [sectionName])
  sectionID = cur.fetchone()[0]
  if sectionID is not None:
    return sectionID
  else: # if no sectionID was found then we need to create the section
    cur.execute("INSERT INTO " + tableName + "(name, order_of_appearance, display_name) VALUES(%s, %s, %s);", (sectionName, 0, sectionName))
    cur.execute("SELECT id FROM " + tableName + " WHERE name = %s;", [sectionName])
    sectionID = cur.fetchone()[0]
    return sectionID


def getGroupID(appName, shapefile, cur):
  group_table = appName + "_ShapefileGroup"
  shapefile_table = appName + "_" + shapefile
  cur.execute("SELECT " + group_table + ".id FROM " + group_table + ", " + shapefile_table + " WHERE " + group_table + ".id=" + shapefile_table + ".group_id")
  ref = cur.fetchone()
  if ref is not None:
    return str(ref[0])
  else:
    print("Could not find a group for the shapefile", shapefile)


def findFilterID(appName, shapefile, key, cur):
  cols = readColumnsFrom(appName, shapefile, cur)
  if len(cols) > 0:
    keyColumn = cols[1]
    cur.execute("SELECT id FROM " + appName + "_" + shapefile + " WHERE " + keyColumn + "::text = %s;", [key])
    ref = cur.fetchone()
    if ref is not None:
      return str(ref[0])
    else: # if cur.fetchone() returns None it means that no matching id was found
      return None
  else: # if readColumnsFrom() returns an empty list it means that nothing was found in the database for the shapefile name we read from snuggetFile
    print("No shapefile with the name", shapefile, "appears to have been loaded.")
    print("If the shapefile exists, you may still need to run the migration and loading steps - see the 'Load some data' section of the readme file.")
    return None




def findAllFilterIDs(appName, shapefile, cur):
  ids = []
  cur.execute("SELECT id FROM " + appName + "_" + shapefile + ";")
  for row in cur.fetchall():
    ids.extend(row)
  return ids




def getSnuggetID(appName, sectionID, subsectionID, filterColumn, filterID, cur):
  cur.execute(
    'SELECT MAX(id) FROM ' + appName + '_snugget WHERE section_id = %s AND sub_section_id = %s AND "' + filterColumn + '" = %s;',
    (sectionID, subsectionID, filterID)
  )
  snuggetID = cur.fetchone()[0]
  if snuggetID is not None:
    return str(snuggetID)
  else:
    return None




def checkForSnugget(appName, sectionID, subsectionID, filterColumn, filterID, cur):
  try:
    snuggetID = getSnuggetID(appName, sectionID, subsectionID, filterColumn, filterID, cur)
    cur.execute("SELECT content FROM " + appName + "_textsnugget WHERE snugget_ptr_id = %s;", [snuggetID])
    return cur.fetchone()[0]
  except: # if nothing came back from the DB, just return None rather than failing
    return None



def removeOldSnugget(appName, sectionID, subsectionID, filterColumn, filterID, cur):
  snuggetID = getSnuggetID(appName, sectionID, subsectionID, filterColumn, filterID, cur)
  if snuggetID is not None:
    cur.execute("DELETE FROM " + appName + "_textsnugget WHERE snugget_ptr_id = %s;", [snuggetID])
    cur.execute("DELETE FROM " + appName + "_snugget WHERE id = %s;", [snuggetID])
    print("Replacing snugget", snuggetID)





# Check with the user about overwriting existing snuggets, giving them the options to:
# either quit and check what's going on, or say "yes to all" and not get prompted again.
def askUserAboutOverwriting(row, oldSnugget, oldSnuggets, snuggetFile, overwriteAlreadySet):
  if overwriteAlreadySet: # if it's already set, then don't do anything else
    return True
  else:
    if oldSnugget is not None:
      print("In shapefile ", repr(row["shapefile"]), " there is already a snugget defined for section " , repr(row["section"]), ", subsection ", repr(row["subsection"]), ", intensity ", repr(row["lookup_value"]), " with the following text content:", sep="")
      print(oldSnugget)
    elif oldSnuggets != []:
      print("In shapefile ", repr(row["shapefile"]), " there are existing snuggets for section" , repr(row["section"]), ", subsection ", repr(row["subsection"]), " with the following text content:", sep="")
      for snugget in oldSnuggets:
        print(snugget)
    else:
      # if no existing snuggets were found, then we neither need to ask the user this time nor change overwriteAll in the calling function
      return overwriteAlreadySet

    print("Please enter one of the following letters to choose how to proceed:")
    print("R: Replace the existing snugget[s] with the new value loaded from", snuggetFile, " and ask again for the next one.")
    print("A: replace All without asking again.")
    print("Q: Quit so you can edit", snuggetFile, "and/or check the snuggets in the Django admin panel and try again.")
    response = ""
    while response not in ["A", "R", "Q"]:
      response = input(">> ").upper()

    if response == "Q":
      exit(0)
    elif response == "A":
      return True
    elif response == "R":
      return False




# drop-in replacement for built-in csv.DictReader() function with .xlsx files
# from https://gist.github.com/mdellavo/853413
# then heavily adapted because the original didn't actually work
def XLSXDictReader(f):
  book  = openpyxl.reader.excel.load_workbook(f)
  sheet = book.get_active_sheet()
  langs = []

  rows = 1
  for row in sheet.iter_rows():
    rows = rows + 1
  cols = 1
  for col in sheet.iter_cols():
    cols = cols + 1

  headers = dict( (i, sheet.cell(row=1, column=i).value) for i in range(1, cols) )
  for header in headers.values():
  	if header.startswith('text-'):
  		langs.append(header.split('-')[1])
  print("Found snugget texts in the following language[s]:", langs, "\n")

  def item(i, j):
    if sheet.cell(row=i, column=j).value == None:
      return (sheet.cell(row=1, column=j).value, '')
    else:
      return (sheet.cell(row=1, column=j).value, str(sheet.cell(row=i, column=j).value))

  return (dict(item(i,j) for j in range(1, cols)) for i in range(2, rows))





if __name__ == "__main__":
  main()
